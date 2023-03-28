import uuid
from typing import List, NamedTuple, Any

import openpyxl.cell.cell
from PyQt6 import QtGui
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QLineEdit, QDialog, QSpinBox, QDateEdit, \
    QPushButton, QFileDialog, QMessageBox

import pandas as pd
from openpyxl.reader.excel import load_workbook

from service import service


class StudentData(NamedTuple):
    user_id: str
    # TODO


class LineEdit(QLineEdit):
    def __init__(self, parent=None):
        QLineEdit.__init__(self, parent=parent)
        self.input_data = self.text()
        self.setPlaceholderText('2342-234234')

    def focusInEvent(self, event):
        self.setInputMask('0000-000000')

    def focusOutEvent(self, event):
        self.input_data = self.text()
        if self.input_data.strip() == "-":
            self.setInputMask('')


class CreateStudent(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QtGui.QIcon("assets/main.ico"))
        self.setWindowTitle("Деканат.Плюс")
        self.setStyleSheet("background-color: #141414")
        self.setFixedSize(850, 710)

        left_column = QVBoxLayout()
        right_column = QVBoxLayout()
        get_student_info = QHBoxLayout()
        get_student_info.setContentsMargins(0, 50, 0, 0)

        title_dialog = QLabel("Создание нового студента")
        title_dialog.setStyleSheet("font-size: 30px; font-weight: semi-bold; color: #FFFFFF;")
        title_dialog.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        self.last_n = QLineEdit()
        self.last_n.setPlaceholderText("Фамилия")
        self.last_n.setFixedSize(250, 45)
        self.last_n.setClearButtonEnabled(True)
        self.last_n.setStyleSheet("color: #FFFFFF; border: 1px solid #FFFFFF; border-radius: 5px; font-size: 14px; "
                                  "padding: 0 0 0 10px; margin-right: 5px;")
        self.first_n = QLineEdit()
        self.first_n.setPlaceholderText("Имя")
        self.first_n.setFixedSize(250, 45)
        self.first_n.setClearButtonEnabled(True)
        self.first_n.setStyleSheet("color: #FFFFFF; border: 1px solid #FFFFFF; border-radius: 5px; font-size: 14px; "
                                   "padding: 0 0 0 10px; margin-right: 5px;")
        self.middle_n = QLineEdit()
        self.middle_n.setPlaceholderText("Отчество")
        self.middle_n.setFixedSize(250, 45)
        self.middle_n.setClearButtonEnabled(True)
        self.middle_n.setStyleSheet("color: #FFFFFF; border: 1px solid #FFFFFF; border-radius: 5px; font-size: 14px; "
                                    "padding: 0 0 0 10px;")
        name_form_layout = QHBoxLayout()
        name_form_layout.addStretch()
        name_form_layout.addWidget(self.last_n)
        name_form_layout.addWidget(self.first_n)
        name_form_layout.addWidget(self.middle_n)
        name_form_layout.addStretch()
        name_form_layout.setContentsMargins(0, 50, 0, 0)


        title_form_edu = QLabel("Форма обучения:")
        title_form_edu.setStyleSheet("font-size: 15px; color: #FFFFFF;")
        self.choice_form_edu = QComboBox()
        self.choice_form_edu.addItems(["Бакалавриат", "Магистратура", "Специалитет"])
        self.choice_form_edu.setFixedSize(200, 30)
        self.choice_form_edu.setStyleSheet("background-color: #424242; color: #FFFFFF; border-radius: 5px;"
                                           "font-size: 14px; padding: 0 0 0 10px;")
        form_edu_layout = QHBoxLayout()
        form_edu_layout.addWidget(title_form_edu)
        form_edu_layout.addWidget(self.choice_form_edu)


        title_number_course = QLabel("Курс обучения:")
        title_number_course.setStyleSheet("font-size: 15px; color: #FFFFFF;")
        self.number_course = QSpinBox()
        self.number_course.setMaximum(5)
        self.number_course.setMinimum(1)
        self.number_course.setFixedSize(100, 30)
        self.number_course.setStyleSheet("background-color: #424242; color: #FFFFFF; border-radius: 5px;"
                                         "font-size: 14px; padding: 0 0 0 10px;")
        number_course_layout = QHBoxLayout()
        number_course_layout.addWidget(title_number_course)
        number_course_layout.addWidget(self.number_course)

        left_column.addLayout(form_edu_layout)
        left_column.addSpacing(10)
        right_column.addLayout(number_course_layout)
        right_column.addSpacing(10)

        get_student_info.addStretch()
        get_student_info.addLayout(left_column)
        get_student_info.addSpacing(160)
        get_student_info.addLayout(right_column)
        get_student_info.addStretch()


        title_group_name = QLabel("Направление:")
        title_group_name.setStyleSheet("font-size: 15px; color: #FFFFFF;")
        self.choice_group_name = QComboBox()
        self.choice_group_name.addItems(["ПМИ", "ФИТ", "ИТХ", "ИТС", "КМБ", "МММ"])
        self.choice_group_name.setFixedSize(200, 30)
        self.choice_group_name.setStyleSheet("background-color: #424242; color: #FFFFFF; border-radius: 5px;"
                                             "font-size: 14px; padding: 0 0 0 10px;")
        group_name_layout = QHBoxLayout()
        group_name_layout.addWidget(title_group_name)
        group_name_layout.addWidget(self.choice_group_name)


        title_number_group = QLabel("Номер группы:")
        title_number_group.setStyleSheet("font-size: 15px; color: #FFFFFF;")
        self.number_group = QSpinBox()
        self.number_group.setMaximum(10)
        self.number_group.setMinimum(1)
        self.number_group.setFixedSize(100, 30)
        self.number_group.setStyleSheet("background-color: #424242; color: #FFFFFF; border-radius: 5px;"
                                        "font-size: 14px; padding: 0 0 0 10px;")
        number_group_layout = QHBoxLayout()
        number_group_layout.addWidget(title_number_group)
        number_group_layout.addWidget(self.number_group)

        left_column.addLayout(group_name_layout)
        left_column.addSpacing(10)
        right_column.addLayout(number_group_layout)
        right_column.addSpacing(10)

        get_student_info.addLayout(left_column)
        get_student_info.addLayout(right_column)


        title_birthday = QLabel("Дата рождения:")
        title_birthday.setStyleSheet("font-size: 15px; color: #FFFFFF;")
        self.birthday = QDateEdit(calendarPopup=True)
        self.birthday.setFixedSize(200, 30)
        self.birthday.setStyleSheet("background-color: #424242; color: #A2A2A2; border-radius: 5px;"
                                    "font-size: 14px; padding: 0 0 0 10px; font-weight: medium;")
        birthday_layout = QHBoxLayout()
        birthday_layout.addWidget(title_birthday)
        birthday_layout.addWidget(self.birthday)


        title_passport_id = QLabel("Данные паспорта:")
        title_passport_id.setStyleSheet("font-size: 15px; color: #FFFFFF;")
        self.passport_id = LineEdit()
        self.passport_id.setFixedSize(130, 30)
        self.passport_id.setStyleSheet("color: #FFFFFF; border: 1px solid #FFFFFF; border-radius: 5px; "
                                       "font-size: 14px; padding: 0 0 0 10px; margin-right: 5px;")
        passport_id_layout = QHBoxLayout()
        passport_id_layout.addWidget(title_passport_id)
        passport_id_layout.addWidget(self.passport_id)

        left_column.addLayout(birthday_layout)
        left_column.addSpacing(10)
        right_column.addLayout(passport_id_layout)
        right_column.addSpacing(10)

        get_student_info.addLayout(left_column)
        get_student_info.addLayout(right_column)


        login_student_title = QLabel("Логин")
        login_student_title.setStyleSheet("font-size: 13px; color: #FFFFFF;")
        self.login_student = QLineEdit()
        self.login_student.setPlaceholderText("Логин")
        self.login_student.setFixedSize(375, 45)
        self.login_student.setText(f"student{str(uuid.uuid4())[:8]}")
        self.login_student.setDisabled(True)
        self.login_student.setStyleSheet("color: #FFFFFF; border: 1px solid #FFFFFF; border-radius: 5px; "
                                         "font-size: 14px; padding: 0 0 0 10px; margin-right: 5px;")

        login_input_layout = QVBoxLayout()
        login_input_layout.addStretch()
        login_input_layout.addWidget(login_student_title, alignment=Qt.AlignmentFlag.AlignLeft)
        login_input_layout.addWidget(self.login_student)
        login_input_layout.addStretch()


        password_student_title = QLabel("Пароль")
        password_student_title.setStyleSheet("font-size: 13px; color: #FFFFFF;")
        self.password_student = QLineEdit()
        self.password_student.setFixedSize(375, 45)
        self.password_student.setText(f"{str(uuid.uuid4())[:3]}{str(uuid.uuid4())[:3]}{str(uuid.uuid4())[:4]}")
        self.password_student.setDisabled(True)
        self.password_student.setStyleSheet("color: #FFFFFF; border: 1px solid #FFFFFF; border-radius: 5px; "
                                            "font-size: 14px; padding: 0 0 0 10px; margin-right: 5px;")
        password_input_layout = QVBoxLayout()
        password_input_layout.addStretch()
        password_input_layout.addWidget(password_student_title, alignment=Qt.AlignmentFlag.AlignLeft)
        password_input_layout.addWidget(self.password_student)
        password_input_layout.addStretch()

        auth_form_layout = QHBoxLayout()
        auth_form_layout.addStretch()
        auth_form_layout.addLayout(login_input_layout)
        auth_form_layout.addLayout(password_input_layout)
        auth_form_layout.addStretch()
        auth_form_layout.setContentsMargins(0, 25, 0, 0)


        download_template_button = QPushButton("Скачать шаблон")
        download_template_button.setFixedSize(250, 40)
        download_template_button.clicked.connect(self.save_file)
        download_template_button.setStyleSheet("background-color: #424242; border: none; border-radius: 5px; "
                                               "font-size: 16px; color: #FFFFFF")
        attach_file_button = QPushButton("Загрузить файл")
        attach_file_button.setFixedSize(250, 40)
        attach_file_button.clicked.connect(self.open_file)
        attach_file_button.setStyleSheet("background-color: #424242; border: none; border-radius: 5px; "
                                         "font-size: 16px; color: #FFFFFF")
        save_data_button = QPushButton("Сохранить")
        save_data_button.setFixedSize(250, 40)
        save_data_button.clicked.connect(self.save_data)
        save_data_button.setStyleSheet("background-color: #424242; border: none; border-radius: 5px; "
                                       "font-size: 16px; color: #FFFFFF")
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()
        buttons_layout.addWidget(download_template_button)
        buttons_layout.addWidget(attach_file_button)
        buttons_layout.addWidget(save_data_button)
        buttons_layout.addStretch()
        buttons_layout.setContentsMargins(0, 35, 0, 0)

        instruction_text = QLabel("Чтобы загрузить студентов из xlsx файла, необходимо скачать шаблон\nи заполнить все"
                                  " поля таблицы, не удаляя заголовок. Затем загрузить шаблон.")
        instruction_text.setStyleSheet("font-size: 15px; color: #FFFFFF;")
        instruction_text.setContentsMargins(0, 20, 0, 0)
        instruction_text.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        dialog_layout = QVBoxLayout()
        dialog_layout.addStretch()
        dialog_layout.addWidget(title_dialog)
        dialog_layout.addLayout(name_form_layout)
        dialog_layout.addLayout(get_student_info)
        dialog_layout.addLayout(auth_form_layout)
        dialog_layout.addLayout(buttons_layout)
        dialog_layout.addWidget(instruction_text)
        dialog_layout.addStretch()
        dialog_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.setLayout(dialog_layout)

    def save_file(self):
        try:
            file_path = QFileDialog.getSaveFileName(self, "Сохранить шаблон", "template.xlsx", "Excel Files (*.xls *.xlsx)")
            writer = pd.ExcelWriter(file_path[0], engine='xlsxwriter')
            df = pd.DataFrame({
                'Фамилия':         ['Ворочаев'],
                'Имя':             ['Артем'],
                'Отчество':        ['Васильевич'],
                'Форма обучения':  ['Бакалавриат'],
                'Курс обучения':   [2],
                'Направление':     ['ПМИ'],
                'Номер группы':    [6],
                'Дата рождения':   ['22.01.2004'],
                'Данные паспорта': ['5171-345657']
            })
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.close()

        except Exception as err:
            print(err)

    def open_file(self):
        try:
            file_path = QFileDialog.getOpenFileName(self, "Выбрать файл", "", "Excel Files (*.xls *.xlsx)")
            wb = load_workbook(file_path[0], read_only=True)
            ws = wb['Sheet1']

            list_students_data = []

            for row in range(1, ws.max_row + 1):
                if ws.cell(row, 1).value == "Фамилия":
                    continue
                student = [
                    ws.cell(row, 1).value,
                    ws.cell(row, 2).value,
                    ws.cell(row, 3).value,
                    ws.cell(row, 4).value,
                    ws.cell(row, 5).value,
                    ws.cell(row, 6).value,
                    ws.cell(row, 7).value,
                    ws.cell(row, 8).value,
                    ws.cell(row, 9).value,
                    f"student{str(uuid.uuid4())[:8]}",
                    f"{str(uuid.uuid4())[:3]}{str(uuid.uuid4())[:3]}{str(uuid.uuid4())[:4]}",
                ]
                validate_status = self.validate_data_xlsx(student)
                if not validate_status:
                    self.alert_message("Ошибка чтения файла", "Возможно, вы неверно указали данные в таблице.")
                    return
                list_students_data.append(student)

            res = service.set_users_student_pull(list_students_data)
            if not res:
                self.alert_message("Ошибка чтения файла", "Возможно, вы неверно указали данные в таблице.")
                return

            self.close()

        except Exception as err:
            print(err)

    def save_data(self) -> None:
        user_data = [
            self.last_n.text().strip(), self.first_n.text().strip(), self.middle_n.text().strip(),
            self.choice_form_edu.currentText(), self.number_course.text(),
            self.choice_group_name.currentText(), self.number_group.text(),
            self.birthday.text(), self.passport_id.text(),
            self.login_student.text(), self.password_student.text()
        ]

        if not self.first_n.text().strip() or not self.last_n.text().strip() or not self.middle_n.text().strip() or not self.passport_id.text():
            self.alert_message("Заполните все поля!", None)
            return

        res = service.set_user_student(user_data)
        if not res:
            self.alert_message("Произошла ошибка, попробуйте еще раз!", None)
            return
        self.close()


    @staticmethod
    def validate_data_xlsx(data: List) -> bool:
        if not data:
            return False

        for index, value in enumerate(data):
            if value is None and index != 2:
                return False
            elif value is None and index == 2:
                data[index] = ""

            if index == 7:
                birthday_data = value.split(".")
                if len(birthday_data) != 3 or len(birthday_data[0]) != 2 or len(birthday_data[1]) != 2 or len(birthday_data[2]) != 4:
                    return False

            if index == 8:
                passportID_data = value.split("-")
                if len(passportID_data) != 2 or len(passportID_data[0]) != 4 or len(passportID_data[1]) != 6:
                    return False
        return True


    @staticmethod
    def alert_message(info: str, more_info: Any) -> None:
        alert = QMessageBox()
        alert.setText(info)
        if more_info is not None:
            alert.setDetailedText(more_info)
        alert.setWindowIcon(QtGui.QIcon("assets/error.ico"))
        alert.setWindowTitle("Что-то пошло не так")
        alert.setStandardButtons(QMessageBox.StandardButton.Close)
        alert.setStyleSheet("color: #141414; font-size: 14px; font-weight: semi-bold;")
        alert.exec()